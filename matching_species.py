from openpyxl import load_workbook

# Function to read species names from the text file
def read_species_from_txt(file_path):
    with open(file_path, 'r') as file:
        species_list = [line.strip() for line in file.readlines()]
    return species_list

# Function to check for matching species in the Excel file
def find_matching_species(txt_species_list, excel_file_path, output_file):
    # Load the workbook and the active sheet
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # Open the output file for writing
    with open(output_file, 'w') as output:
        match_count = 0
        # Iterate over the rows of the Excel sheet, starting from row 1
        for row in sheet.iter_rows(min_row=1, max_col=4, values_only=True):
            excel_species_name = str(row[3])  # Fourth column is index 3
            if excel_species_name in txt_species_list:
                output.write(f"{excel_species_name}\n")
                match_count += 1

    return match_count

# File paths
txt_file_path = '/ocean/projects/bio200049p/zjiang2/Files/spring24/cactus241names.txt'  # Path to your text file
excel_file_path = '/ocean/projects/bio200049p/zjiang2/Files/spring24/470names.xlsx'  # Path to your Excel file
output_file_path = '/ocean/projects/bio200049p/zjiang2/Files/spring24/matched_species.txt'  # Output file for matched species

# Main process
txt_species = read_species_from_txt(txt_file_path)
match_count = find_matching_species(txt_species, excel_file_path, output_file_path)

print(f"Total matched species: {match_count}")
